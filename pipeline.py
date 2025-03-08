import openpyxl.styles
import yfinance as yf
import pandas as pd
import openpyxl
import numpy as np
import matplotlib.pyplot as plt
from scipy.stats import pearsonr

#Given a set of tickers, create an excel workbook with a separate sheet for each stock. 
class ASXPipeline:
    #Initialise excel workbook if it does not already exist
    def __init__(self, tickers, workbookPath, initialiseWorkbook=False):
        self.tickers = tickers
        self.workbookPath = workbookPath
        if initialiseWorkbook:
            self.createWorkbook()

    def createWorkbook(self):
        self.tickers.append("^AXJO")
        with pd.ExcelWriter(self.workbookPath, engine="openpyxl", mode="w") as writer:
            for tickerName in self.tickers:
                print(f"Fetching data for {tickerName}...")

                # Fetch 1 year stock data
                ticker = yf.Ticker(tickerName)
                historical_data = ticker.history(period="1y") 

                historical_data.index = historical_data.index.tz_localize(None)
                historical_data.reset_index(inplace=True)
                historical_data["Date"] = pd.to_datetime(historical_data["Date"]).dt.date  # Removes time

                #Add returns column
                historical_data["Returns"] = historical_data["Close"].pct_change()

                # Write to a new sheet in the Excel file
                historical_data.to_excel(writer, sheet_name=tickerName, index=False)

    
        wb = openpyxl.load_workbook(self.workbookPath)
        

        date_style = openpyxl.styles.NamedStyle(name="datetime", number_format="YYYY-MM-DD")

        for tickerName in self.tickers:
            ws = wb[tickerName]
            for cell in ws["A"][1:]:
                cell.style = date_style

       
        wb.save(self.workbookPath)
        print(f"Excel file '{self.workbookPath}, monitoring tickers: {self.tickers}")

    #Plot data trends for given tickers
   
    def plot(self, tickers, xAxis="Date", yAxis="Open"):
        plt.figure(figsize=(12, 6))

        for ticker in tickers:
            df = pd.read_excel(self.workbookPath, sheet_name=ticker, engine="openpyxl")
            df["Date"] = pd.to_datetime(df["Date"])

            if xAxis not in df.columns or yAxis not in df.columns:
                raise ValueError(f"{xAxis} or {yAxis} is an invalid input for {ticker}")

            xData = df[xAxis].to_numpy()
            yData = df[yAxis].to_numpy()

            plt.plot(xData, yData, label=ticker)  # Plot each ticker

        plt.xlabel(xAxis)
        plt.ylabel(yAxis)
        plt.title(f"{yAxis} vs {xAxis} for {[t for t in tickers]}")
        plt.legend()
        plt.grid(True)
        plt.show()

    
    def statSummary(self, ticker):
        df = pd.read_excel(self.workbookPath, sheet_name=ticker, engine="openpyxl")
        stats = {
            "Mean": df["Close"].mean(),
            "Median": df["Close"].median(),
            "Variance": df["Close"].var(),
            "Standard Deviation": df["Close"].std(),
            "Skewness": df["Close"].skew(),
            "Kurtosis": df["Close"].kurt(),
            "Five Number Summary": df["Close"].describe(percentiles=[0.25, 0.5, 0.75]).to_dict()
        }

        return stats
    
    #Returns both a correlation and p-value for closing values of tickers
    def correlate(self, ticker1, ticker2):
        df1 = pd.read_excel(self.workbookPath, sheet_name=ticker1, engine="openpyxl")[["Date", "Close"]]
        df2 = pd.read_excel(self.workbookPath, sheet_name=ticker2, engine="openpyxl")[["Date", "Close"]]

        df1["Date"] = pd.to_datetime(df1["Date"])
        df2["Date"] = pd.to_datetime(df2["Date"])

        merged_df = pd.merge(df1, df2, on="Date", suffixes=(f"_{ticker1}", f"_{ticker2}"))

        # Compute correlation coefficient & p-value
        r, p_value = pearsonr(merged_df[f"Close_{ticker1}"], merged_df[f"Close_{ticker2}"])

        return f"Correlation: {r:.2f}, p-value: {p_value:.5f}"


    #plot volatility => simply std deviations of returns within some given time period
    def plot_volatility(self, tickers, window=20):
        plt.figure(figsize=(12, 6))

        for ticker in tickers:
            df = pd.read_excel(self.workbookPath, sheet_name=ticker, engine="openpyxl")
            df["Date"] = pd.to_datetime(df["Date"])

            df["Returns"] = df["Close"].pct_change()


            df["Volatility"] = df["Returns"].rolling(window=window).std(ddof=1)

            dates = df["Date"].to_numpy()
            volatility = df["Volatility"].to_numpy()

            plt.plot(dates, volatility, label=f"{ticker}")

        plt.xlabel("Date")
        plt.ylabel("Volatility (Std Dev of Returns)")
        plt.title(f"Rolling {window}-Day Volatility for {tickers}")
        plt.legend()
        plt.grid(True)
        plt.show()

    #Compare the performance of a given ticker against ASX200
    def plotReturnsAgainstASX200(self, ticker):
        asx200 = pd.read_excel(self.workbookPath, sheet_name="^AXJO", engine="openpyxl")
        stock = pd.read_excel(self.workbookPath, sheet_name=ticker, engine="openpyxl")
        asx200["Date"] = pd.to_datetime(asx200["Date"])
        stock["Date"] = pd.to_datetime(stock["Date"])

       
        merged_df = pd.merge(stock[["Date", "Returns"]], 
                            asx200[["Date", "Returns"]], 
                            on="Date", 
                            suffixes=("", "_ASX200"))

 
        merged_df.dropna(inplace=True)

        
        dates = merged_df["Date"].to_numpy()
        stockReturns = merged_df["Returns"].to_numpy()
        asxReturns = merged_df["Returns_ASX200"].to_numpy()

        # Plot
        plt.figure(figsize=(12, 6))  
        plt.plot(dates, stockReturns, label=f"{ticker} Returns", color='blue')
        plt.plot(dates, asxReturns, label="ASX200 Returns", color='red', linestyle="dashed")

        plt.xlabel("Date")
        plt.ylabel("Daily Returns (%)")
        plt.title(f"{ticker} vs ASX200 Daily Returns")
        plt.legend()
        plt.grid(True)
        plt.show()


def createTestFile():
    tickerName = "AAPL"
    ticker = yf.Ticker(tickerName)
    historical_data = ticker.history(period="1y") 
    historical_data.index = historical_data.index.tz_localize(None)
    historical_data.reset_index(inplace=True) 
    historical_data["Date"] = pd.to_datetime(historical_data["Date"])
    print(historical_data["Date"])
    fileName = "output1.xlsx"
    with pd.ExcelWriter(fileName, engine="openpyxl", mode="w") as writer:
        historical_data.to_excel(writer, sheet_name=tickerName)

    wb = openpyxl.load_workbook(fileName)
    ws = wb[tickerName]


    date_style = openpyxl.styles.NamedStyle(name="datetime", number_format="YYYY-MM-DD")

    for cell in ws["A"][1:]:  
        cell.style = date_style
    wb.save(fileName)



def testPlot(xAxis="Date", yAxis="Open"):
    df = pd.read_excel("output1.xlsx", sheet_name="AAPL", engine="openpyxl")
    df["Date"] = pd.to_datetime(df["Date"])
    if xAxis not in df.columns:
        raise ValueError(f"{xAxis} is an invalid input")
    if yAxis not in df.columns:
        raise ValueError(f"{yAxis} is an invalid input")
    
    xData = df[xAxis].to_numpy()
    yData = df[yAxis].to_numpy()
    plt.plot(xData, yData)
    plt.show()
